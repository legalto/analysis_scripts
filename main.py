import os
import argparse
import openpyxl
import re

def excel_to_list(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))

def sort_cities_by_pop(cities):
    # sort cities by population in second param. Reverse=True makes it in descending order from largest
    return(sorted(cities, key = lambda x: x[1], reverse=True))

def remove_duplicates(cities):
    initial_len = len(cities)
    totals = {}
    for k,v in cities:
        totals[k] = totals.get(k,0) + v
    print("Trimmed: ", initial_len - len(totals))
    return map(list, totals.items())

def main(args):

    # open cities dataset
    cities_list = excel_to_list(args.cities)
    cities_header = cities_list[0]  # read header for writing later
    cities_list = cities_list[1:]   # strip header from dataset
    cities_list = remove_duplicates(cities_list)

    # make sure it is sorted
    cities_list = sort_cities_by_pop(cities_list)

    # reading dataset_list
    dataset_list = excel_to_list(args.dataset)
    dataset_header = dataset_list[0] # read header
    dataset_list = dataset_list[1:] # strip header

    # open a new workbook (our output)
    myworkbook = openpyxl.Workbook()

    match_count = 0
    # iterate through each city
    for city, population in cities_list:
        first_city_match = True

        # for each row in the dataset, check if the city is in the first element of that row
        for keyword, volume, kd, cpc in dataset_list:
            if re.search(r'\b'+city+r'\b', keyword):
                #city is in keyword

                # check if this is the first time we found a match to write sheet and header in myworkbook
                if first_city_match == True:
                    ws_write = myworkbook.create_sheet(city)
                    ws_write.append(dataset_header)
                    first_city_match = False
                    match_count = match_count + 1
                ws_write = myworkbook.get_sheet_by_name(city)
                #write data since we have a match
                ws_write.append([keyword,volume,kd,cpc])

    # save our workbook
    myworkbook.save(args.output)
    print("City Matches Found: ", match_count)
if __name__ == '__main__':


    # Three arguments cities of interest, dataset, output file name
    # Run: python -c cities.xlsx -d dataset.xlsx -o output.xlsx
    parser = argparse.ArgumentParser(description="Generates tiles for registration")
    parser.add_argument("-c","--cities", required=False, help="path to root directory containing map tiles over multiple dates")
    parser.add_argument("-d","--dataset", required=False, help="path to dataset to generate a truth dataset file")
    parser.add_argument("-o", "--output", required=False, help="path to dataset to generate a truth dataset file")
    args = parser.parse_args()

    main(args)
